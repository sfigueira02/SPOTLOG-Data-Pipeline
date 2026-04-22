import streamlit as st
import pandas as pd
import openpyxl
import plotly.graph_objects as go
import plotly.express as px
from pathlib import Path

EXCEL_PATH  = Path(__file__).parent / "Organized Data_19_11_2025.xlsx"
GM_CSV_PATH = Path(__file__).parent / "google_maps_valores.csv"

VEHICLE_NAMES = {
    37: "Citroen C4 — DIESEL",
    40: "Fiat Panda — MHEV",
    43: "Volkswagen ID.3 — Elétrico",
    46: "Peugeot 3008 — MHEV",
}
ELECTRIC_VEHICLE_ID = 43

# Each corridor = (prefix, label, origin, destination)
# prefix maps to R→route_prefix+"_E" and prefix+"_G"
CORRIDORS = [
    ("AR", "Aveiro  |  Univ. Aveiro → Ílhavo"),
    ("AL", "Aveiro  |  Ílhavo → Univ. Aveiro"),
    ("CR", "Coimbra  |  Mercado de Santiago → Hotel Mélia Ria"),
    ("CL", "Coimbra  |  Hotel Mélia Ria → Mercado de Santiago"),
    ("DR", "Sever/Albergaria  |  Continente Albergaria → CM Sever do Vouga"),
    ("DL", "Sever/Albergaria  |  CM Sever do Vouga → Continente Albergaria"),
]
CORRIDOR_LABELS = {prefix: label for prefix, label in CORRIDORS}
CORRIDOR_PREFIXES = [prefix for prefix, _ in CORRIDORS]

# Metrics
METRICS_COMBUSTION = {
    "CO2 (g)": "co2_g",
    "CO2 (g/km)": "co2_gkm",
    "Consumo (L/100km)": "fuel_l100km",
    "NOx (g)": "nox_g",
    "NOx (mg/km)": "nox_mgkm",
    "HC (g)": "hc_g",
    "HC (mg/km)": "hc_mgkm",
    "CO (g)": "co_g",
    "CO (mg/km)": "co_mgkm",
    "Distância (km)": "dist_km",
    "Tempo de viagem (s)": "travel_time_s",
}
METRICS_ELECTRIC = {
    "Energia (kWh)": "energy_kwh",
    "Energia (kWh/km)": "energy_kwhkm",
    "Distância (km)": "dist_km",
    "Tempo de viagem (s)": "travel_time_s",
}
DEFAULT_COMBUSTION = {"CO2 (g)", "CO2 (g/km)", "Consumo (L/100km)", "NOx (g)", "Distância (km)"}
DEFAULT_ELECTRIC   = {"Energia (kWh)", "Energia (kWh/km)", "Distância (km)"}


@st.cache_data
def load_data() -> pd.DataFrame:
    import datetime
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # Build Trip ID → date map from "Organized Data_No PEMS" (col 1=TripID, col 11=Date)
    trip_dates: dict = {}
    ws_org = wb["Organized Data_No PEMS"]
    for row in ws_org.iter_rows(min_row=2, values_only=True):
        tid = row[1]
        date_val = row[11]
        if tid and date_val and tid not in trip_dates:
            if isinstance(date_val, datetime.datetime):
                trip_dates[tid] = date_val.date()
            elif isinstance(date_val, datetime.date):
                trip_dates[tid] = date_val

    ws = wb["Results"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    trip_counters: dict = {}

    for row in rows[1:]:
        vehicle_id = row[0]
        if not isinstance(vehicle_id, (int, float)):
            continue
        vehicle_id = int(vehicle_id)

        trip_id    = row[1]
        route      = row[2]
        name       = row[3]
        energy_kw  = row[4]
        travel_s   = row[5]
        dist_m     = row[6]
        co2_g      = row[8]
        nox_g      = row[9]
        hc_g       = row[10]
        co_g       = row[11]

        if not isinstance(dist_m, (int, float)) or dist_m == 0 or route is None:
            continue

        dist_km = dist_m / 1000

        key = (vehicle_id, route)
        trip_counters[key] = trip_counters.get(key, 0) + 1
        trip_num = trip_counters[key]

        # Derive corridor prefix (AR, AL, CR, CL, DR, DL)
        corridor = route[:2] if len(route) >= 2 else None
        route_type = route[3] if len(route) > 3 else (route[2] if len(route) > 2 else None)

        energy_kwh    = energy_kw if isinstance(energy_kw, (int, float)) else None
        energy_kwhkm  = energy_kwh / dist_km if energy_kwh is not None else None

        co2_g_val     = co2_g if isinstance(co2_g, (int, float)) else None
        co2_gkm       = co2_g_val / dist_km if co2_g_val else None
        fuel_l100km   = co2_gkm * 0.043103448275862 if co2_gkm else None
        nox_g_val     = nox_g if isinstance(nox_g, (int, float)) else None
        nox_mgkm      = nox_g_val / dist_km * 1000 if nox_g_val else None
        hc_g_val      = hc_g if isinstance(hc_g, (int, float)) else None
        hc_mgkm       = hc_g_val / dist_km * 1000 if hc_g_val else None
        co_g_val      = co_g if isinstance(co_g, (int, float)) else None
        co_mgkm       = co_g_val / dist_km * 1000 if co_g_val else None

        records.append({
            "vehicle_id":   vehicle_id,
            "vehicle_name": VEHICLE_NAMES.get(vehicle_id, f"Veículo {vehicle_id}"),
            "is_electric":  vehicle_id == ELECTRIC_VEHICLE_ID,
            "trip_id":      trip_id,
            "trip_num":     trip_num,
            "date":         trip_dates.get(trip_id),
            "route":        route,
            "route_type":   route_type,
            "corridor":     corridor,
            "name":         name,
            "travel_time_s": travel_s,
            "dist_km":      dist_km,
            "co2_g":        co2_g_val,
            "co2_gkm":      co2_gkm,
            "fuel_l100km":  fuel_l100km,
            "nox_g":        nox_g_val,
            "nox_mgkm":     nox_mgkm,
            "hc_g":         hc_g_val,
            "hc_mgkm":      hc_mgkm,
            "co_g":         co_g_val,
            "co_mgkm":      co_mgkm,
            "energy_kwh":   energy_kwh,
            "energy_kwhkm": energy_kwhkm,
        })

    return pd.DataFrame(records)


def fmt(val, col: str) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "N/D"
    if col == "travel_time_s":
        return f"{int(val)//60}m {int(val)%60}s"
    if col == "dist_km":         return f"{val:.3f} km"
    if col in ("co2_g", "nox_g", "hc_g", "co_g"):
        return f"{val:.4f} g"
    if col == "co2_gkm":         return f"{val:.2f} g/km"
    if col == "fuel_l100km":     return f"{val:.2f} L/100km"
    if col in ("nox_mgkm", "hc_mgkm", "co_mgkm"):
        return f"{val:.4f} mg/km"
    if col == "energy_kwh":      return f"{val:.4f} kWh"
    if col == "energy_kwhkm":    return f"{val:.4f} kWh/km"
    return f"{val:.4f}"


def make_comparison_df(row_eco, row_alt, metrics: dict, label_eco: str, label_alt: str) -> pd.DataFrame:
    data = []

    # Date row at the top
    date_eco = row_eco.get("date") if row_eco is not None else None
    date_alt = row_alt.get("date") if row_alt is not None else None
    data.append({
        "Métrica": "Data",
        label_eco: date_eco.strftime("%d/%m/%Y") if date_eco else "—",
        label_alt: date_alt.strftime("%d/%m/%Y") if date_alt else "—",
        "Δ (Alt − Eco)": "",
        "_color": "",
    })

    for m_label, m_col in metrics.items():
        ve = row_eco.get(m_col) if row_eco is not None else None
        va = row_alt.get(m_col) if row_alt is not None else None

        diff_text = "N/D"
        color_hint = ""
        if ve is not None and va is not None and not pd.isna(ve) and not pd.isna(va):
            diff = va - ve
            pct  = (diff / ve * 100) if ve != 0 else None
            arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "=")
            pct_s = f" ({abs(pct):.1f}%)" if pct is not None else ""
            diff_text = f"{arrow} {abs(diff):.4f}{pct_s}"
            color_hint = "red" if diff > 0 else ("green" if diff < 0 else "")

        data.append({
            "Métrica": m_label,
            label_eco: fmt(ve, m_col),
            label_alt: fmt(va, m_col),
            f"Δ (Alt − Eco)": diff_text,
            "_color": color_hint,
        })
    return pd.DataFrame(data)


def render_comparison(df_cmp: pd.DataFrame):
    display_cols = [c for c in df_cmp.columns if c != "_color"]

    def highlight(row):
        styles = [""] * len(row)
        val = row.iloc[-1]  # last col is the Δ column
        if isinstance(val, str):
            if "▲" in val:
                styles[-1] = "color: #d32f2f; font-weight: bold"
            elif "▼" in val:
                styles[-1] = "color: #2e7d32; font-weight: bold"
        return styles

    st.dataframe(
        df_cmp[display_cols].style.apply(highlight, axis=1),
        use_container_width=True,
        hide_index=True,
    )


def bar_chart_pair(row_eco, row_alt, m_col: str, m_label: str, label_eco: str, label_alt: str):
    vals = [
        {"Rota": label_eco, m_col: row_eco.get(m_col) if row_eco is not None else None},
        {"Rota": label_alt, m_col: row_alt.get(m_col) if row_alt is not None else None},
    ]
    df_p = pd.DataFrame(vals).dropna(subset=[m_col])
    if df_p.empty:
        st.info("Sem dados para este gráfico.")
        return
    colors = ["#1976D2", "#F57C00"]
    fig = px.bar(df_p, x="Rota", y=m_col, color="Rota",
                 color_discrete_sequence=colors,
                 labels={"Rota": "", m_col: m_label},
                 title=m_label, text=m_col)
    fig.update_traces(texttemplate="%{text:.3f}", textposition="outside")
    fig.update_layout(showlegend=False, height=360)
    st.plotly_chart(fig, use_container_width=True)


def all_trips_chart(df_v: pd.DataFrame, corridor: str, m_col: str, m_label: str):
    """Line chart: ECO vs ALT across all trip numbers for a corridor."""
    route_e = f"{corridor}_E"
    route_g = f"{corridor}_G"
    df_e = df_v[df_v["route"] == route_e][["trip_num", m_col]].dropna().rename(columns={m_col: "Eco"})
    df_g = df_v[df_v["route"] == route_g][["trip_num", m_col]].dropna().rename(columns={m_col: "Alternativa"})
    df_merged = df_e.merge(df_g, on="trip_num", how="outer").sort_values("trip_num")
    if df_merged.empty:
        st.info("Sem dados suficientes.")
        return
    fig = go.Figure()
    if "Eco" in df_merged:
        fig.add_trace(go.Scatter(x=df_merged["trip_num"], y=df_merged["Eco"],
                                 mode="lines+markers", name="Eco",
                                 line=dict(color="#1976D2", width=2)))
    if "Alternativa" in df_merged:
        fig.add_trace(go.Scatter(x=df_merged["trip_num"], y=df_merged["Alternativa"],
                                 mode="lines+markers", name="Alternativa",
                                 line=dict(color="#F57C00", width=2)))
    fig.update_layout(
        xaxis_title="Nº da Viagem",
        yaxis_title=m_label,
        title=f"{m_label} — todas as viagens",
        height=360,
        xaxis=dict(tickmode="linear", dtick=1),
    )
    st.plotly_chart(fig, use_container_width=True)


def load_gm_data() -> pd.DataFrame:
    if not GM_CSV_PATH.exists():
        return pd.DataFrame(columns=["veiculo_id","veiculo","corredor","viagem","gm_poupanca_pct","gm_metrica"])
    df = pd.read_csv(GM_CSV_PATH, dtype={"veiculo_id": int, "viagem": int, "gm_poupanca_pct": object})
    # Coerce poupança to float where possible
    df["gm_poupanca_pct"] = pd.to_numeric(df["gm_poupanca_pct"], errors="coerce")
    return df


def save_gm_data(df: pd.DataFrame):
    df.to_csv(GM_CSV_PATH, index=False)


def get_gm_value(df_gm: pd.DataFrame, vehicle_id: int, corridor: str, trip_num: int):
    """Return (pct, metrica) or (None, None) if not set."""
    mask = (
        (df_gm["veiculo_id"] == vehicle_id) &
        (df_gm["corredor"] == corridor) &
        (df_gm["viagem"] == trip_num)
    )
    row = df_gm[mask]
    if row.empty:
        return None, None
    pct = row.iloc[0]["gm_poupanca_pct"]
    met = row.iloc[0].get("gm_metrica", "combustivel")
    return (float(pct) if pd.notna(pct) else None), met


# ─────────────────────────────────────────────────────────────
def main():
    st.set_page_config(page_title="Comparador de Rotas", page_icon="🚗", layout="wide")
    st.title("🚗 Comparador de Rotas — Eco vs Alternativa")

    df = load_data()
    if df.empty:
        st.error("Não foi possível carregar dados.")
        return

    # ── Sidebar ───────────────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ Configurações")

        # 1. Veículo
        vehicles = sorted(df["vehicle_id"].unique())
        vehicle_id = st.selectbox(
            "🚙  Veículo",
            options=vehicles,
            format_func=lambda v: VEHICLE_NAMES.get(v, f"V{v}"),
        )
        is_electric = vehicle_id == ELECTRIC_VEHICLE_ID
        metrics = METRICS_ELECTRIC if is_electric else METRICS_COMBUSTION
        defaults = DEFAULT_ELECTRIC if is_electric else DEFAULT_COMBUSTION

        df_v = df[df["vehicle_id"] == vehicle_id]

        # 2. Corredor (zona + sentido)
        # Only show corridors that have both E and G routes for this vehicle
        available_corridors = []
        for prefix, label in CORRIDORS:
            has_e = not df_v[df_v["route"] == f"{prefix}_E"].empty
            has_g = not df_v[df_v["route"] == f"{prefix}_G"].empty
            if has_e or has_g:
                available_corridors.append((prefix, label))

        corridor_prefix = st.selectbox(
            "🗺️  Corredor",
            options=[p for p, _ in available_corridors],
            format_func=lambda p: CORRIDOR_LABELS.get(p, p),
        )

        route_eco = f"{corridor_prefix}_E"
        route_alt = f"{corridor_prefix}_G"

        # 3. Nº da viagem — intersection of trips available in BOTH routes
        trips_e = set(df_v[df_v["route"] == route_eco]["trip_num"].unique())
        trips_g = set(df_v[df_v["route"] == route_alt]["trip_num"].unique())
        trips_both = sorted(trips_e & trips_g)
        trips_union = sorted(trips_e | trips_g)

        trip_options = trips_both if trips_both else trips_union
        if not trip_options:
            st.warning("Sem viagens disponíveis para este corredor.")
            return

        trip_num = st.selectbox(
            "🔢  Viagem nº",
            options=trip_options,
            format_func=lambda n: f"Viagem {n}",
        )

        st.markdown("---")

        # 4. Métricas
        st.subheader("📊 Métricas")
        selected_metrics = {}
        for label, col in metrics.items():
            if st.checkbox(label, value=(label in defaults)):
                selected_metrics[label] = col

    # ── Get the two rows ─────────────────────────────────────
    row_eco_df = df_v[(df_v["route"] == route_eco) & (df_v["trip_num"] == trip_num)]
    row_alt_df = df_v[(df_v["route"] == route_alt) & (df_v["trip_num"] == trip_num)]

    row_eco = row_eco_df.iloc[0] if not row_eco_df.empty else None
    row_alt = row_alt_df.iloc[0] if not row_alt_df.empty else None

    label_eco = f"{route_eco}  (Viagem {trip_num})"
    label_alt = f"{route_alt}  (Viagem {trip_num})"

    # ── Top info bar ─────────────────────────────────────────
    vname = VEHICLE_NAMES.get(vehicle_id, f"V{vehicle_id}")
    icon  = "⚡" if is_electric else "⛽"
    corr_label = CORRIDOR_LABELS.get(corridor_prefix, corridor_prefix)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Veículo", f"{icon} {vname.split('—')[1].strip()}")
    c2.metric("Corredor", corr_label.split("|")[0].strip())
    c3.metric("Sentido", corr_label.split("|")[1].strip() if "|" in corr_label else "")
    c4.metric("Viagem", f"#{trip_num}")

    st.markdown("---")

    # ── Comparison table (always visible) ────────────────────
    if not selected_metrics:
        st.info("Seleciona pelo menos uma métrica na barra lateral.")
    elif row_eco is None and row_alt is None:
        st.error("Sem dados para este corredor + viagem.")
    else:
        tab_cmp, tab_charts, tab_resumo = st.tabs([
            "⚖️ Comparação", "📊 Gráficos", "📋 Resumo"
        ])

        # ── Tab 1: Comparison table ───────────────────────────
        with tab_cmp:
            date_eco = row_eco.get("date") if row_eco is not None else None
            date_alt = row_alt.get("date") if row_alt is not None else None
            trip_eco_id = int(row_eco["trip_id"]) if row_eco is not None else "—"
            trip_alt_id = int(row_alt["trip_id"]) if row_alt is not None else "—"
            date_eco_str = date_eco.strftime("%d/%m/%Y") if date_eco else "—"
            date_alt_str = date_alt.strftime("%d/%m/%Y") if date_alt else "—"
            st.caption(
                f"🟦 Eco **{route_eco}**: {date_eco_str} (Trip ID {trip_eco_id})   |   "
                f"🟧 Alt **{route_alt}**: {date_alt_str} (Trip ID {trip_alt_id})   |   "
                f"▲ = Alt tem valor mais alto   ▼ = Alt tem valor mais baixo"
            )
            df_cmp = make_comparison_df(row_eco, row_alt, selected_metrics, label_eco, label_alt)
            render_comparison(df_cmp)

        # ── Tab 2: Charts ─────────────────────────────────────
        with tab_charts:
            col_left, col_right = st.columns(2)
            with col_left:
                st.subheader("Esta viagem")
                chart_metric = st.selectbox("Métrica", list(selected_metrics.keys()), key="chart_single")
                m_col = selected_metrics[chart_metric]
                bar_chart_pair(row_eco, row_alt, m_col, chart_metric, label_eco, label_alt)
            with col_right:
                st.subheader("Evolução — todas as viagens")
                chart_metric2 = st.selectbox("Métrica", list(selected_metrics.keys()), key="chart_all")
                m_col2 = selected_metrics[chart_metric2]
                all_trips_chart(df_v, corridor_prefix, m_col2, chart_metric2)

        # ── Tab 3: Summary table ──────────────────────────────
        with tab_resumo:
            st.subheader("Todas as viagens deste corredor")

            def build_summary(route: str, label: str) -> pd.DataFrame:
                df_r = df_v[df_v["route"] == route].copy()
                if df_r.empty:
                    return pd.DataFrame()
                cols_show = ["trip_num", "date"] + [c for c in selected_metrics.values() if c in df_r.columns]
                df_r = df_r[cols_show].copy()
                df_r["date"] = df_r["date"].apply(
                    lambda d: d.strftime("%d/%m/%Y") if d is not None else "—"
                )
                df_r.insert(0, "Rota", label)
                return df_r.rename(columns={"trip_num": "Viagem", "date": "Data"})

            df_sum_e = build_summary(route_eco, "Eco")
            df_sum_g = build_summary(route_alt, "Alternativa")
            df_summary = pd.concat([df_sum_e, df_sum_g], ignore_index=True).sort_values(["Viagem", "Rota"])
            rename_metrics = {v: k for k, v in selected_metrics.items()}
            df_summary = df_summary.rename(columns=rename_metrics)
            st.dataframe(df_summary.reset_index(drop=True), use_container_width=True, hide_index=True)



if __name__ == "__main__":
    main()
